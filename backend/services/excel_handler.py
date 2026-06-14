import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from typing import Dict, List, Any
import re

class ExcelHandler:
    """
    Handles Excel file reading and writing operations
    """
    
    # System template columns
    SYSTEM_COLUMNS = [
        "رقم السطر",
        "اسم المرسل",
        "الراسل الفرعي",
        "محافظة الراسل",
        "منطقة الراسل",
        "عنوان الراسل",
        "أرقام الراسل",
        "اسم المستلم",
        "العنوان",
        "المحافظة",
        "المنطقة",
        "رقم الموبايل",
        "رقم الهاتف",
        "رقم البوليصة",
        "السعر",
        "كود العميل",
        "الخدمة",
        "نوع الطلب",
        "نوع التحصيل",
        "نوع السعر",
        "الوزن",
        "عدد القطع",
        "فتح الطرد",
        "رقم المرجع",
        "الملاحظات"
    ]
    
    def read_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Read Excel file and return structured data
        """
        try:
            # Read with openpyxl to preserve formatting
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get headers
            headers = []
            for cell in ws[1]:
                value = cell.value if cell.value else ""
                headers.append(str(value).strip())
            
            # Get data rows
            data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                row_data = {}
                has_data = False
                
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        value = cell.value
                        # Convert to string if not None
                        if value is not None:
                            row_data[headers[col_idx]] = str(value).strip()
                            if str(value).strip():
                                has_data = True
                        else:
                            row_data[headers[col_idx]] = ""
                
                # Skip empty rows
                if has_data:
                    data.append(row_data)
            
            return {
                "columns": headers,
                "data": data,
                "row_count": len(data)
            }
        
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def write_excel(self, file_path: str, columns: List[str], data: List[Dict]) -> None:
        """
        Write data to Excel file maintaining system template
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "الشحنات"
            
            # Write headers with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col_idx, header in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            data_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, column in enumerate(columns, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = row_data.get(column, "")
                    
                    # Convert None and empty to empty string
                    if value is None or value == "":
                        cell.value = ""
                    else:
                        cell.value = value
                    
                    cell.alignment = data_alignment
            
            # Adjust column widths
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18
            
            # Set row height for header
            ws.row_dimensions[1].height = 30
            
            # Save workbook
            wb.save(file_path)
        
        except Exception as e:
            raise Exception(f"Error writing Excel file: {str(e)}")
    
    def get_auto_mappings(self, customer_columns: List[str], system_columns: List[str]) -> Dict[str, str]:
        """
        Auto-detect column mappings based on similarity
        """
        mappings = {}
        
        # Define mapping rules
        mapping_rules = {
            "اسم المستلم": ["customer_name", "name", "اسم", "المستقبل", "اسم المستقبل", "اسم العميل"],
            "العنوان": ["address", "location", "العنوان", "موقع", "المكان"],
            "المحافظة": ["province", "governorate", "محافظة", "الولاية", "State"],
            "المنطقة": ["area", "district", "منطقة", "القسم", "Region"],
            "رقم الموبايل": ["mobile", "phone", "موبايل", "رقم الجوال", "رقم الموبايل", "Mobile", "Phone"],
            "رقم الهاتف": ["phone2", "alternate_phone", "الهاتف", "Phone2", "Tel"],
            "رقم البوليصة": ["waybill", "tracking", "bill_number", "رقم البوليصة", "رقم التتبع", "Waybill"],
            "السعر": ["price", "amount", "total", "السعر", "المبلغ", "Price", "Total"],
        }
        
        # Try to match columns
        used_customer_cols = set()
        
        for system_col in system_columns:
            if system_col in mapping_rules:
                for customer_col in customer_columns:
                    if customer_col in used_customer_cols:
                        continue
                    
                    customer_col_lower = customer_col.lower()
                    
                    # Exact match
                    if customer_col_lower == system_col.lower():
                        mappings[system_col] = customer_col
                        used_customer_cols.add(customer_col)
                        break
                    
                    # Partial match
                    for rule in mapping_rules[system_col]:
                        if rule.lower() in customer_col_lower or customer_col_lower in rule.lower():
                            if system_col not in mappings:
                                mappings[system_col] = customer_col
                                used_customer_cols.add(customer_col)
                            break
        
        return mappings
